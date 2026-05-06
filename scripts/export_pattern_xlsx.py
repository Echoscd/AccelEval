#!/usr/bin/env python3
"""
export_pattern_xlsx.py — flatten the per-model agent_pattern_analysis.json
files into a single multi-sheet xlsx for manual exploration.

Sheets:
  README           : column legend + provenance
  Patterns_Long    : one row per (model, task, pattern) — the flat raw join
  Strategies       : one row per (model, task) — strategy + bottleneck + speedup
  New_Candidates   : one row per (model, task, candidate) proposed by analyzer
  Pattern_x_Model  : coverage matrix, rows=PAT-id, cols=model, value=# tasks attributed
  Pattern_x_Task   : coverage matrix, rows=PAT-id, cols=task,  value=# models attributed

Usage:
  python3 scripts/export_pattern_xlsx.py
  python3 scripts/export_pattern_xlsx.py --out runs/reports/pattern_breakdown_<ts>.xlsx
"""
from __future__ import annotations

import argparse
import glob
import json
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent

RUNS_L3 = {
    "Gemini 3.1 Pro":  "gemini-3.1-pro-preview-openrouter_l3_20260421_1712",
    "Claude Opus 4.6": "claude-opus-4.6-openrouter_l3_20260421_1905",
    "GPT-5.4":         "openai/gpt-5.4_l3_20260421_1905",
    "Qwen 3.6 Plus":   "qwen/qwen3.6-plus_l3_20260421_1905",
    "Kimi K2.5":       "kimi-k2.5-openrouter_l3_20260421_1905",
    "GLM 5.1":         "glm-5.1-openrouter_l3_20260421_1905",
    "DeepSeek V3.2":   "deepseek-v3.2-openrouter_l3_20260421_1905",
    "DeepSeek V4 Pro": "deepseek-v4-pro-openrouter_l3_20260424_1706",
}


def _truncate(s: str | None, n: int = 32000) -> str:
    if s is None:
        return ""
    s = str(s)
    if len(s) <= n:
        return s
    return s[: n - 20] + " …(truncated)"


def load_pattern_catalog():
    """Return ordered (PAT-id -> name) and (PAT-id -> category) from KB."""
    kb = json.loads((ROOT / "Library" / "knowledge_data" / "patterns.json").read_text())
    by_id = {}
    cat = {}
    for p in sorted(kb["patterns"], key=lambda x: x["id"]):
        by_id[p["id"]] = p["name"]
        cat[p["id"]] = p["category"]
    return by_id, cat


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=None,
                    help="Output xlsx path (default: runs/reports/pattern_breakdown_<ts>.xlsx)")
    args = ap.parse_args()

    out_path = Path(args.out) if args.out else (
        ROOT / "runs" / "reports" / f"pattern_breakdown_{time.strftime('%Y%m%d_%H%M')}.xlsx"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)

    pat_name, pat_cat = load_pattern_catalog()
    pat_ids = list(pat_name.keys())  # 43 patterns, PAT-001…PAT-043

    long_rows = []
    strategy_rows = []
    new_rows = []
    # coverage[pid][model] = set of tasks
    coverage_model: dict[str, dict[str, set]] = {pid: {m: set() for m in RUNS_L3} for pid in pat_ids}
    coverage_task: dict[str, dict[str, set]] = {pid: {} for pid in pat_ids}

    n_pairs = 0
    for model_name, run_subdir in RUNS_L3.items():
        p = ROOT / "runs" / run_subdir / "agent_pattern_analysis.json"
        if not p.exists():
            print(f"[skip] {model_name}: missing {p}")
            continue
        data = json.loads(p.read_text())
        for task_id, entry in data.items():
            n_pairs += 1
            meta = entry.get("_meta") or {}
            speedup = meta.get("speedup_e2e", 0.0) or 0.0

            ps_list = entry.get("pattern_summaries") or []
            nc_list = entry.get("new_candidates") or []

            strategy_rows.append([
                model_name, task_id,
                round(speedup, 3),
                len(ps_list),
                len(nc_list),
                _truncate(entry.get("strategy_summary"), 4000),
                _truncate(entry.get("bottleneck_analysis"), 4000),
                meta.get("auto_matched_count", 0),
                ",".join(meta.get("auto_matched_pattern_ids") or []),
                meta.get("analyzed_at", ""),
            ])

            for ps in ps_list:
                pid = ps.get("pattern_id", "")
                if pid in coverage_model:
                    coverage_model[pid][model_name].add(task_id)
                    coverage_task[pid].setdefault(task_id, set()).add(model_name)

                long_rows.append([
                    model_name, task_id,
                    round(speedup, 3),
                    pid,
                    ps.get("pattern_name") or pat_name.get(pid, ""),
                    pat_cat.get(pid, ""),
                    ps.get("source", ""),
                    _truncate(ps.get("target"), 1500),
                    _truncate(ps.get("method"), 1500),
                    _truncate(ps.get("intensity_note"), 200),
                    _truncate(ps.get("code_evidence"), 8000),
                ])

            for nc in nc_list:
                new_rows.append([
                    model_name, task_id,
                    round(speedup, 3),
                    nc.get("estimated_impact", ""),
                    _truncate(nc.get("raw_description"), 2000),
                    _truncate(nc.get("mechanism_hypothesis"), 2000),
                    _truncate(nc.get("code_snippet"), 8000),
                ])

    print(f"[load] {n_pairs} (model, task) pairs  |  "
          f"{len(long_rows)} pattern rows  |  {len(new_rows)} new candidates")

    # ----- write xlsx -----
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    wrap = Alignment(wrap_text=True, vertical="top")

    def write_header(ws, headers):
        ws.append(headers)
        for c, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c)
            cell.font = bold
            cell.fill = header_fill
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions  # set later after data

    # README
    ws = wb.active
    ws.title = "README"
    ws.append(["AccelEval — pattern decomposition raw data"])
    ws.append([f"Generated:  {time.strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([f"Source:     runs/<model>/agent_pattern_analysis.json (8 models, medium passes)"])
    ws.append([f"Analyzer:   gemini-3.1-pro-preview-openrouter (43-pattern KB)"])
    ws.append([f"Coverage:   {n_pairs} (model, task) pairs analyzed"])
    ws.append([])
    ws.append(["Sheets:"])
    sheet_doc = [
        ("Patterns_Long",   "one row per (model, task, attributed pattern) — the flat raw join, easy to pivot in Excel"),
        ("Strategies",      "one row per (model, task) — overall GPU strategy summary + bottleneck + counts"),
        ("New_Candidates",  "techniques the analyzer flagged as novel (not in the 43-pattern catalog)"),
        ("Pattern_x_Model", "coverage matrix: rows = PAT-id, columns = model, cell = # tasks where that model used that pattern"),
        ("Pattern_x_Task",  "coverage matrix: rows = PAT-id, columns = task,  cell = # models that used that pattern on that task"),
    ]
    for n, desc in sheet_doc:
        ws.append([n, desc])
    ws.append([])
    ws.append(["source column meaning:"])
    ws.append(["  auto_detected", "matched by static keyword / regex scan over the .cu source"])
    ws.append(["  manual_check",  "verified by the LLM analyzer reading the code (e.g. coalescing, AoS->SoA layout)"])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110

    # Patterns_Long
    ws = wb.create_sheet("Patterns_Long")
    headers = ["model", "task", "speedup_e2e", "pattern_id", "pattern_name", "category",
               "source", "target", "method", "intensity_note", "code_evidence"]
    write_header(ws, headers)
    for row in long_rows:
        ws.append(row)
    widths = [16, 28, 12, 12, 36, 24, 14, 36, 36, 22, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(ord('A') + i - 1)].width = w
    ws.auto_filter.ref = ws.dimensions

    # Strategies
    ws = wb.create_sheet("Strategies")
    headers = ["model", "task", "speedup_e2e", "n_patterns", "n_new",
               "strategy_summary", "bottleneck_analysis",
               "auto_matched_count", "auto_matched_ids", "analyzed_at"]
    write_header(ws, headers)
    for row in strategy_rows:
        ws.append(row)
    widths = [16, 28, 12, 11, 8, 70, 50, 16, 36, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(ord('A') + i - 1)].width = w
    ws.auto_filter.ref = ws.dimensions

    # New_Candidates
    ws = wb.create_sheet("New_Candidates")
    headers = ["model", "task", "speedup_e2e", "estimated_impact",
               "raw_description", "mechanism_hypothesis", "code_snippet"]
    write_header(ws, headers)
    for row in new_rows:
        ws.append(row)
    widths = [16, 28, 12, 14, 60, 50, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(ord('A') + i - 1)].width = w
    ws.auto_filter.ref = ws.dimensions

    # Pattern_x_Model
    ws = wb.create_sheet("Pattern_x_Model")
    models = list(RUNS_L3.keys())
    headers = ["pattern_id", "pattern_name", "category"] + models + ["row_total_tasks"]
    write_header(ws, headers)
    for pid in pat_ids:
        row = [pid, pat_name[pid], pat_cat[pid]]
        per_model = [len(coverage_model[pid][m]) for m in models]
        union_tasks = set()
        for m in models:
            union_tasks |= coverage_model[pid][m]
        row.extend(per_model)
        row.append(len(union_tasks))
        ws.append(row)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 26
    for i, _ in enumerate(models, 4):
        ws.column_dimensions[chr(ord('A') + i - 1)].width = 16
    ws.auto_filter.ref = ws.dimensions

    # Pattern_x_Task
    # Build the union of tasks observed (across all models)
    all_tasks = sorted({t for pid in pat_ids for t in coverage_task[pid].keys()})
    ws = wb.create_sheet("Pattern_x_Task")
    headers = ["pattern_id", "pattern_name"] + all_tasks + ["row_total_models_seen"]
    write_header(ws, headers)
    for pid in pat_ids:
        row = [pid, pat_name[pid]]
        sum_models = 0
        for t in all_tasks:
            n = len(coverage_task[pid].get(t, set()))
            row.append(n)
            sum_models += n
        row.append(sum_models)
        ws.append(row)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    ws.auto_filter.ref = ws.dimensions

    wb.save(out_path)
    print(f"[done] -> {out_path}  ({out_path.stat().st_size/1024:.1f} KB)")


if __name__ == "__main__":
    main()

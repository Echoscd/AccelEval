#!/usr/bin/env python3
"""
analyze_pattern_impact.py — rank the 43 catalog patterns by their
estimated contribution to speedup, using the per-(model, task) breakdown
JSONs produced by analyze_medium_solutions.py.

Three orthogonal signals (from the methodology proposed earlier):

  1. Winner coverage     f_winner(P) = #{task t : per-task winner used P} / 42
                         f_random(P) = #{(m, t) : m used P on t} / N_pairs
                         "selectivity" = f_winner / f_random
                         (high selectivity ⇒ winners disproportionately reach for P)

  2. Within-task lift    For each task t, partition models that passed t into
                         U = used P, V = didn't use P. If both nonempty,
                            lift_t(P) = mean log s_m,t (m in U) − mean log s_m,t (m in V)
                         LIFT(P) = exp( mean over t of lift_t(P) )  — reported as ×
                         (controls for task difficulty; the cleanest marginal signal)

  3. Bundles             Jaccard similarity over the same set of (model, task)
                         attributions. Patterns that almost always co-occur
                         (Jaccard >= 0.6) are flagged as bundles — single-attribution
                         is unreliable for them.

Output: pattern_impact_<ts>.xlsx with sheets:
  README           — column legend + methodology
  Impact_Ranked    — 43 patterns ranked by LIFT, with all signals + bundle flag
  Pattern_x_Pattern— Jaccard co-occurrence matrix
  Bundles          — connected components above the Jaccard threshold

Usage:
  python3 scripts/analyze_pattern_impact.py
"""
from __future__ import annotations

import argparse
import json
import math
import time
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent

RUNS_L3 = {
    "Gemini 3.1 Pro":  "gemini-3.1-pro-preview-openrouter_l3_20260421_1712",
    "Claude Opus 4.6": "claude-opus-4.6-openrouter_l3_20260421_1905",
    "GPT-5.4":         "openai/gpt-5.4_l3_20260421_1905",
    "Qwen 3.6 Plus":   "qwen/qwen3.6-plus_l3_20260421_1905",
    "Kimi K2.5":       "kimi-k2.5-openrouter_l3_20260421_1905",
    "GLM 5.1":         "glm-5.1-openrouter_l3_20260421_1905",
    "DeepSeek V3.2":   "deepseek-v3.2-openrouter_l3_20260421_1905",
    "DeepSeek V4 Pro": "deepseek-v4-pro-openrouter_l3_20260424_1706",
}

JACCARD_THRESHOLD = 0.6


def load_catalog():
    kb = json.loads((ROOT / "Library" / "knowledge_data" / "patterns.json").read_text())
    patterns = sorted(kb["patterns"], key=lambda x: x["id"])
    by_id = {p["id"]: p for p in patterns}
    return by_id


def load_observations():
    """Yield {model, task, speedup, patterns:set[PAT-id]} per (model, task) pair."""
    obs = []
    for model, run_subdir in RUNS_L3.items():
        p = ROOT / "runs" / run_subdir / "agent_pattern_analysis.json"
        if not p.exists():
            print(f"[skip] {model}: missing analysis file")
            continue
        data = json.loads(p.read_text())
        for task, entry in data.items():
            speedup = ((entry.get("_meta") or {}).get("speedup_e2e") or 0.0)
            if speedup <= 0:
                continue  # log() needs positive
            pids = set()
            for ps in entry.get("pattern_summaries") or []:
                pid = ps.get("pattern_id")
                if pid and pid.startswith("PAT-"):
                    pids.add(pid)
            obs.append({
                "model": model,
                "task": task,
                "speedup": speedup,
                "patterns": pids,
            })
    return obs


def compute_winner_per_task(obs):
    """Per task, return the model with max speedup."""
    by_task = defaultdict(list)
    for o in obs:
        by_task[o["task"]].append(o)
    winners = {}
    for t, lst in by_task.items():
        winners[t] = max(lst, key=lambda x: x["speedup"])
    return winners, by_task


def compute_lift(by_task: dict, pid: str) -> tuple[float, int, int]:
    """Within-task log-lift averaged across tasks for pattern pid.
    Returns (LIFT_x, n_tasks_with_both, n_paired_models).
    """
    diffs = []
    paired_n = 0
    for t, lst in by_task.items():
        with_p = [math.log(o["speedup"]) for o in lst if pid in o["patterns"]]
        without_p = [math.log(o["speedup"]) for o in lst if pid not in o["patterns"]]
        if not with_p or not without_p:
            continue
        diffs.append(sum(with_p) / len(with_p) - sum(without_p) / len(without_p))
        paired_n += min(len(with_p), len(without_p))
    if not diffs:
        return float("nan"), 0, 0
    return math.exp(sum(diffs) / len(diffs)), len(diffs), paired_n


def compute_jaccard_matrix(obs, pat_ids):
    """For each pattern P, the set S_P of (model, task) attributions.
    Jaccard(P, Q) = |S_P ∩ S_Q| / |S_P ∪ S_Q|.
    """
    sets: dict[str, set] = {pid: set() for pid in pat_ids}
    for o in obs:
        for pid in o["patterns"]:
            if pid in sets:
                sets[pid].add((o["model"], o["task"]))
    n = len(pat_ids)
    J = [[0.0] * n for _ in range(n)]
    for i, pi in enumerate(pat_ids):
        for j, pj in enumerate(pat_ids):
            a, b = sets[pi], sets[pj]
            if not a or not b:
                J[i][j] = 0.0
                continue
            J[i][j] = len(a & b) / len(a | b)
    return J, sets


def find_bundles(pat_ids, J, threshold=JACCARD_THRESHOLD):
    """Connected components in the threshold-graph of Jaccard(P, Q) >= threshold."""
    n = len(pat_ids)
    parent = list(range(n))
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x
    def union(x, y):
        x, y = find(x), find(y)
        if x != y:
            parent[x] = y
    for i in range(n):
        for j in range(i + 1, n):
            if J[i][j] >= threshold:
                union(i, j)
    groups: dict[int, list[str]] = defaultdict(list)
    for i, pid in enumerate(pat_ids):
        groups[find(i)].append(pid)
    bundles = [g for g in groups.values() if len(g) > 1]
    return bundles


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=None)
    ap.add_argument("--jaccard", type=float, default=JACCARD_THRESHOLD,
                    help=f"Jaccard threshold for bundling (default {JACCARD_THRESHOLD})")
    args = ap.parse_args()

    out_path = Path(args.out) if args.out else (
        ROOT / "runs" / "reports" / f"pattern_impact_{time.strftime('%Y%m%d_%H%M')}.xlsx"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)

    catalog = load_catalog()
    pat_ids = sorted(catalog.keys())
    obs = load_observations()
    winners, by_task = compute_winner_per_task(obs)
    n_tasks = len(by_task)
    n_pairs = len(obs)
    print(f"[load] {n_pairs} pairs over {n_tasks} tasks")

    # Coverage counts
    f_winner = {pid: 0 for pid in pat_ids}
    f_random = {pid: 0 for pid in pat_ids}
    for o in obs:
        for pid in o["patterns"]:
            if pid in f_random:
                f_random[pid] += 1
    for t, w in winners.items():
        for pid in w["patterns"]:
            if pid in f_winner:
                f_winner[pid] += 1

    # Jaccard + bundles
    J, sets = compute_jaccard_matrix(obs, pat_ids)
    bundles = find_bundles(pat_ids, J, threshold=args.jaccard)
    bundle_of = {}
    for k, b in enumerate(bundles):
        for pid in b:
            bundle_of[pid] = k + 1  # bundle id 1, 2, …

    # Lift + assemble rows
    rows = []
    for pid in pat_ids:
        meta = catalog[pid]
        lift_x, n_tasks_lift, paired_n = compute_lift(by_task, pid)
        fw = f_winner[pid] / n_tasks
        fr = f_random[pid] / n_pairs
        sel = fw / fr if fr > 0 else float("inf")
        rows.append({
            "pattern_id": pid,
            "pattern_name": meta["name"],
            "category": meta["category"],
            "auto_detectable": meta["auto_detectable"],
            "n_pairs_attrib": f_random[pid],
            "n_winner_tasks": f_winner[pid],
            "f_winner": round(fw, 4),
            "f_random": round(fr, 4),
            "selectivity": round(sel, 3) if not math.isinf(sel) else "inf",
            "lift_x": round(lift_x, 3) if not math.isnan(lift_x) else None,
            "n_tasks_in_lift": n_tasks_lift,
            "paired_n": paired_n,
            "bundle_id": bundle_of.get(pid, ""),
        })

    # Sort rows for the ranked sheet: NaN lift goes to the bottom
    def lift_key(r):
        v = r["lift_x"]
        return (v is None or v == "" or (isinstance(v, float) and math.isnan(v)),
                -(v if isinstance(v, (int, float)) and not math.isnan(v) else 0))
    ranked = sorted(rows, key=lift_key)

    # ---- write xlsx ----
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import ColorScaleRule

    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")

    def write_header(ws, headers):
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = bold
            cell.fill = header_fill
        ws.freeze_panes = "A2"

    # README
    ws = wb.active
    ws.title = "README"
    ws.append(["AccelEval — pattern impact ranking"])
    ws.append([f"Generated:  {time.strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([f"Source:     runs/<model>/agent_pattern_analysis.json"])
    ws.append([f"Coverage:   {n_pairs} (model, task) pairs over {n_tasks} tasks (medium passes only)"])
    ws.append([])
    ws.append(["Methodology:"])
    ws.append(["  f_winner(P)", "= #{task t : per-task winner used P} / N_tasks"])
    ws.append(["  f_random(P)", "= #{(m, t) : m used P on t} / N_pairs"])
    ws.append(["  selectivity",  "= f_winner / f_random  (>1 ⇒ winners disproportionately reach for P)"])
    ws.append(["  LIFT (×)",     "= exp( mean over tasks t of [ mean_logsp(used) − mean_logsp(unused) ] )"])
    ws.append([],)
    ws.append([f"  bundle_id", f"non-empty when this pattern co-occurs with another at Jaccard ≥ {args.jaccard}"])
    ws.append([f"            ", "(single-attribution lift is unreliable for these — see Bundles sheet)"])
    ws.append([])
    ws.append(["Reading guide:"])
    ws.append(["  • LIFT > 1.5×  & selectivity > 1.2 ⇒ winning pattern (rare-and-effective)"])
    ws.append(["  • LIFT ≈ 1×     & f_random > 0.7    ⇒ universal table-stakes"])
    ws.append(["  • LIFT < 1×     & paired_n large    ⇒ counter-indicator (slow when used)"])
    ws.append(["  • paired_n small (< 5)              ⇒ low-confidence; ignore the lift"])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95

    # Impact_Ranked
    ws = wb.create_sheet("Impact_Ranked")
    headers = ["pattern_id", "pattern_name", "category", "auto_det",
               "n_pairs", "n_winner",
               "f_winner", "f_random", "selectivity",
               "LIFT_x", "n_tasks_lift", "paired_n", "bundle_id"]
    write_header(ws, headers)
    for r in ranked:
        ws.append([
            r["pattern_id"], r["pattern_name"], r["category"],
            "Y" if r["auto_detectable"] else "N",
            r["n_pairs_attrib"], r["n_winner_tasks"],
            r["f_winner"], r["f_random"], r["selectivity"],
            r["lift_x"], r["n_tasks_in_lift"], r["paired_n"], r["bundle_id"],
        ])
    widths = [12, 40, 26, 8, 8, 8, 9, 9, 11, 9, 11, 9, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(ord('A') + i - 1)].width = w
    ws.auto_filter.ref = ws.dimensions
    # Color scales: LIFT_x col J (10), selectivity col I (9), f_winner col G (7)
    last = ws.max_row
    rule = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="num", mid_value=1, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    )
    ws.conditional_formatting.add(f"J2:J{last}", rule)
    ws.conditional_formatting.add(f"I2:I{last}", rule)
    rule_g = ColorScaleRule(
        start_type="min", start_color="FFFFFF",
        end_type="max", end_color="63BE7B",
    )
    ws.conditional_formatting.add(f"G2:H{last}", rule_g)

    # Pattern_x_Pattern
    ws = wb.create_sheet("Pattern_x_Pattern")
    headers = ["pattern_id", "pattern_name"] + pat_ids
    write_header(ws, headers)
    for i, pid in enumerate(pat_ids):
        row = [pid, catalog[pid]["name"]] + [round(J[i][j], 3) for j in range(len(pat_ids))]
        ws.append(row)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    for k in range(len(pat_ids)):
        col = ws.cell(row=1, column=3 + k).column_letter
        ws.column_dimensions[col].width = 10
    last = ws.max_row
    cf_rule = ColorScaleRule(
        start_type="num", start_value=0, start_color="FFFFFF",
        mid_type="num",   mid_value=0.3,  mid_color="FFEB84",
        end_type="num",   end_value=1,    end_color="63BE7B",
    )
    last_col_letter = ws.cell(row=1, column=2 + len(pat_ids)).column_letter
    ws.conditional_formatting.add(f"C2:{last_col_letter}{last}", cf_rule)
    ws.auto_filter.ref = f"A1:{last_col_letter}{last}"

    # Bundles
    ws = wb.create_sheet("Bundles")
    write_header(ws, ["bundle_id", "size", "pattern_ids", "pattern_names"])
    if bundles:
        for k, b in enumerate(bundles, 1):
            ids_str = ", ".join(b)
            names_str = " + ".join(catalog[pid]["name"] for pid in b)
            ws.append([k, len(b), ids_str, names_str])
    else:
        ws.append([f"(no bundles at Jaccard >= {args.jaccard})"])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 80

    wb.save(out_path)
    print(f"[done] -> {out_path}  ({out_path.stat().st_size/1024:.1f} KB)")
    print()
    # Print a short top/bottom summary to console
    valid = [r for r in ranked if isinstance(r["lift_x"], (int, float)) and r["paired_n"] >= 5]
    if valid:
        top = sorted(valid, key=lambda r: r["lift_x"], reverse=True)[:5]
        print("[top-5 LIFT, paired_n ≥ 5]")
        for r in top:
            print(f"  {r['pattern_id']:7} {r['pattern_name'][:42]:42} "
                  f"LIFT={r['lift_x']:.2f}× sel={r['selectivity']} "
                  f"f_win={r['f_winner']} (n={r['paired_n']})")
        bot = sorted(valid, key=lambda r: r["lift_x"])[:5]
        print()
        print("[bottom-5 LIFT, paired_n ≥ 5]")
        for r in bot:
            print(f"  {r['pattern_id']:7} {r['pattern_name'][:42]:42} "
                  f"LIFT={r['lift_x']:.2f}× sel={r['selectivity']} "
                  f"f_win={r['f_winner']} (n={r['paired_n']})")
        print()
        if bundles:
            print(f"[bundles ≥ Jaccard {args.jaccard}] ({len(bundles)} found):")
            for k, b in enumerate(bundles, 1):
                print(f"  bundle {k}: {' + '.join(catalog[pid]['name'] for pid in b)}")


if __name__ == "__main__":
    main()

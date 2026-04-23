#!/usr/bin/env python3
"""Export consolidated XLSX with small + medium results side by side."""
import json, os, glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

RUNS = {
    "Gemini-3.1-Pro":   "gemini-3.1-pro-preview-openrouter_l3_20260421_1712",
    "Claude-Opus-4.6":  "claude-opus-4.6-openrouter_l3_20260421_1905",
    "GPT-5.4":          "openai/gpt-5.4_l3_20260421_1905",
    "Qwen-3.6-Plus":    "qwen/qwen3.6-plus_l3_20260421_1905",
    "Kimi-K2.5":        "kimi-k2.5-openrouter_l3_20260421_1905",
    "GLM-5.1":          "glm-5.1-openrouter_l3_20260421_1905",
    "DeepSeek-V3.2":    "deepseek-v3.2-openrouter_l3_20260421_1905",
}
MODELS = list(RUNS.keys())


def load_eval(run_name):
    """Merge all eval_results_*.json for a run. Later files override earlier."""
    files = sorted(glob.glob(os.path.join(ROOT, "runs", run_name, "eval_results_*.json")))
    if not files: return {}
    merged = {}
    for p in files:
        with open(p) as f:
            d = json.load(f)
        for k, r in d.items():
            tid = r.get("task_id")
            b = r.get("benchmark") or {}
            size = b.get("size_name", "")
            # Key by (task_id, size) so same task on different sizes coexist
            # Default: if no size, treat as compile_fail (overwrite by task_id)
            if size:
                merged[(tid, size)] = r
            else:
                # Only set compile_fail entries if no successful benchmark exists
                if (tid, "small") not in merged and (tid, "medium") not in merged and (tid, "large") not in merged:
                    merged[(tid, "")] = r
                # Also always keep if same (tid, "") was earlier — later overrides
                if (tid, "") in merged:
                    merged[(tid, "")] = r
    # Convert back to flat dict keyed by sample_key, with task_id preserved in values
    out = {}
    for (tid, size), r in merged.items():
        out[f"{tid}_{size}"] = r
    return out


def classify(r, size):
    """For a given eval result, return (status, speedup) for the given size."""
    b = r.get("benchmark") or {}
    r_size = b.get("size_name")
    if r_size == size:
        if r.get("correct"):
            sp = b.get("speedup_e2e")
            return ("pass", sp if sp and sp > 0 else None)
        if r.get("compiled"):
            return ("wrong", None)
        return ("fail", None)
    if r_size is None and not r.get("compiled"):
        return ("fail", None)
    return (None, None)  # evaluated for different size, treat as missing


def stats_for(model_data, size):
    """Return stats for one model on one size. Only counts tasks that were
    actually evaluated at that size (has a benchmark with matching size_name)."""
    tasks = {}  # tid -> result dict (latest per tid)
    for r in model_data.values():
        tid = r.get("task_id")
        b = r.get("benchmark") or {}
        if b.get("size_name") == size:
            tasks[tid] = r
    total = len(tasks)
    n_pass = n_wrong = n_fail = 0
    sps = []
    for r in tasks.values():
        if r.get("correct"):
            n_pass += 1
            sp = (r.get("benchmark") or {}).get("speedup_e2e")
            if sp and sp > 0: sps.append(sp)
        elif r.get("compiled"):
            n_wrong += 1
        else:
            n_fail += 1
    sps.sort()
    med = sps[len(sps)//2] if sps else 0
    avg = sum(sps)/len(sps) if sps else 0
    return {
        "tasks": total,
        "pass": n_pass, "wrong": n_wrong, "fail": n_fail,
        "pass_rate": n_pass/total if total else 0,
        "compile_rate": (n_pass+n_wrong)/total if total else 0,
        "median": med, "avg": avg,
        "fast1": sum(1 for s in sps if s>=1)/total if total else 0,
        "fast5": sum(1 for s in sps if s>=5)/total if total else 0,
        "fast10": sum(1 for s in sps if s>=10)/total if total else 0,
    }


def main():
    wb = Workbook()

    all_data = {m: load_eval(rn) for m, rn in RUNS.items()}

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    pass_fill = PatternFill("solid", fgColor="D9EAD3")
    wrong_fill = PatternFill("solid", fgColor="FCE5CD")
    fail_fill = PatternFill("solid", fgColor="F4CCCC")

    # ── Sheet 1: Leaderboard (small + medium side by side) ──
    ws = wb.active
    ws.title = "Leaderboard"
    header1 = ["Rank", "Model",
               "small Tasks", "small Pass%", "small Median", "small fast@10",
               "medium Tasks", "medium Pass%", "medium Median", "medium fast@10"]
    ws.append(header1)
    for c in range(1, len(header1)+1):
        ws.cell(row=1, column=c).fill = header_fill
        ws.cell(row=1, column=c).font = header_font
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    rows = []
    for m in MODELS:
        s_stats = stats_for(all_data[m], "small")
        m_stats = stats_for(all_data[m], "medium")
        rows.append((m, s_stats, m_stats))
    # Sort by medium pass_rate (primary) then small
    rows.sort(key=lambda r: (-r[2]["pass_rate"], -r[1]["pass_rate"]))

    for rank, (m, s, mm) in enumerate(rows, 1):
        ws.append([rank, m,
                   s["tasks"], s["pass_rate"], s["median"], s["fast10"],
                   mm["tasks"], mm["pass_rate"], mm["median"], mm["fast10"]])
        r = ws.max_row
        for col in [4,6,8,10]:
            ws.cell(row=r, column=col).number_format = "0.0%"
        for col in [5,9]:
            ws.cell(row=r, column=col).number_format = "0.0"

    widths = [6, 20, 11, 11, 12, 13, 13, 12, 13, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    # ── Sheet 2 & 3: Speedup matrix per size ──
    for size in ["small", "medium"]:
        ws2 = wb.create_sheet(f"{size} matrix")
        h = ["Task", "Category", "Diff"] + MODELS
        ws2.append(h)
        for c in range(1, len(h)+1):
            ws2.cell(row=1, column=c).fill = header_fill
            ws2.cell(row=1, column=c).font = header_font
            ws2.cell(row=1, column=c).alignment = Alignment(horizontal="center")

        # Collect all task ids
        all_tasks = set()
        for md in all_data.values():
            for r in md.values():
                all_tasks.add(r.get("task_id"))
        all_tasks.discard(None)
        task_meta = {}
        for tid in all_tasks:
            tj = os.path.join(ROOT, "tasks", tid, "task.json")
            if os.path.exists(tj):
                with open(tj) as f: d = json.load(f)
                task_meta[tid] = (d.get("category",""), d.get("difficulty",0))
            else:
                task_meta[tid] = ("REMOVED", 0)

        for tid in sorted(all_tasks):
            cat, diff = task_meta[tid]
            row_vals = [tid, cat, "★"*diff]
            ws2.append(row_vals)
            r = ws2.max_row
            for i, m in enumerate(MODELS):
                c = 4 + i
                md = all_data[m]
                # find this task's result
                matching = [rr for rr in md.values() if rr.get("task_id")==tid]
                if not matching:
                    ws2.cell(row=r, column=c).value = "-"
                    continue
                rr = matching[0]
                status, sp = classify(rr, size)
                if status == "pass":
                    ws2.cell(row=r, column=c).value = sp if sp else ""
                    ws2.cell(row=r, column=c).fill = pass_fill
                    ws2.cell(row=r, column=c).number_format = "0.00"
                elif status == "wrong":
                    ws2.cell(row=r, column=c).value = "wrong"
                    ws2.cell(row=r, column=c).fill = wrong_fill
                elif status == "fail":
                    ws2.cell(row=r, column=c).value = "fail"
                    ws2.cell(row=r, column=c).fill = fail_fill
                else:
                    ws2.cell(row=r, column=c).value = "-"
        ws2.column_dimensions["A"].width = 34
        ws2.column_dimensions["B"].width = 22
        ws2.column_dimensions["C"].width = 6
        for i in range(len(MODELS)):
            ws2.column_dimensions[get_column_letter(4+i)].width = 14

    out = os.path.join(ROOT, "runs/reports/benchmark_20260422.xlsx")
    wb.save(out)
    print(f"Saved: {out}")

    # Also print leaderboard to stdout
    print("\n=== MEDIUM LEADERBOARD ===")
    print(f"{'Rank':<5}{'Model':<18}{'Pass':>7}{'Median':>9}{'fast@10':>9}")
    for i, (m, s, mm) in enumerate(rows, 1):
        med_s = f"{mm['median']:.1f}x"
        print(f"{i:<5}{m:<18}{mm['pass_rate']:>6.1%} {med_s:>8} {mm['fast10']:>8.1%}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Build an xlsx comparing Stage 1 (no guidance) vs Stage 2 (with guidance) at medium.

Uses GEOMETRIC MEAN as the headline metric (correct for multiplicative speedups).

Two main sheets:
  1. "Summary"  — per-model: pass count, GM speedup, fast@10x, paired-task GM lift.
                  Includes DeepSeek V4 Pro (Stage 1 only — no Stage 2 was generated).
  2. "Per-task" — task x model matrix; each cell shows S1, S2, S2/S1 ratio with
                  color coding.

Usage: python3 scripts/export_s1_vs_s2_xlsx.py
"""
import json
import math
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent

S1 = {
    "Gemini 3.1 Pro":  "gemini-3.1-pro-preview-openrouter_l3_20260421_1712",
    "Claude Opus 4.6": "claude-opus-4.6-openrouter_l3_20260421_1905",
    "GPT-5.4":         "openai/gpt-5.4_l3_20260421_1905",
    "Qwen 3.6 Plus":   "qwen/qwen3.6-plus_l3_20260421_1905",
    "Kimi K2.5":       "kimi-k2.5-openrouter_l3_20260421_1905",
    "GLM 5.1":         "glm-5.1-openrouter_l3_20260421_1905",
    "DeepSeek V3.2":   "deepseek-v3.2-openrouter_l3_20260421_1905",
    "DeepSeek V4 Pro": "deepseek-v4-pro-openrouter_l3_20260424_1706",  # baseline only
}
S2 = {
    "Gemini 3.1 Pro":  "gemini-3.1-pro-preview-openrouter_l3s2_20260424_1534",
    "Claude Opus 4.6": "claude-opus-4.6-openrouter_l3s2_20260424_1534",
    "GPT-5.4":         "openai/gpt-5.4_l3s2_20260424_1534",
    "Qwen 3.6 Plus":   "qwen/qwen3.6-plus_l3s2_20260424_1534",
    "Kimi K2.5":       "kimi-k2.5-openrouter_l3s2_20260424_1534",
    "GLM 5.1":         "glm-5.1-openrouter_l3s2_20260424_1534",
    "DeepSeek V3.2":   "deepseek-v3.2-openrouter_l3s2_20260424_1534",
    "DeepSeek V4 Pro": "deepseek-v4-pro-openrouter_l3s2_20260427_1312",
}
SIZE = "medium"


def gmean(xs):
    xs = [x for x in xs if x and x > 0]
    if not xs:
        return 0.0
    return math.exp(sum(math.log(x) for x in xs) / len(xs))


def latest_per_task(rd: str, size: str = SIZE) -> dict:
    d = ROOT / "runs" / rd
    lat = {}
    for ev in d.glob("eval_results_*.json"):
        mt = ev.stat().st_mtime
        try:
            doc = json.loads(ev.read_text())
            res = doc.get("results", doc)
            for k, r in res.items():
                if not isinstance(r, dict):
                    continue
                bm = r.get("benchmark") or {}
                if bm.get("size_name") != size:
                    continue
                t = r.get("task_id") or k.rsplit("_sample_", 1)[0]
                if t not in lat or mt > lat[t][0]:
                    lat[t] = (mt, r)
        except Exception:
            continue
    return {t: r for t, (_, r) in lat.items()}


def passed(r):
    return bool(r and r.get("compiled") and r.get("correct"))


def speedup(r):
    if not passed(r):
        return None
    return ((r.get("benchmark") or {}).get("speedup_e2e")) or None


def main():
    ref = ROOT / "runs" / "claude-opus-4.6-openrouter_l3s2_20260424_1534"
    tasks = sorted(d.name for d in ref.iterdir() if d.is_dir())
    n = len(tasks)

    s1_data = {m: latest_per_task(d) for m, d in S1.items()}
    s2_data = {m: latest_per_task(d) for m, d in S2.items()}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)
    italic = Font(italic=True)
    center = Alignment(horizontal="center")
    blue  = PatternFill("solid", fgColor="D9E1F2")
    gf    = PatternFill("solid", fgColor="C6EFCE")
    lgf   = PatternFill("solid", fgColor="DDF2D5")
    rf    = PatternFill("solid", fgColor="FFC7CE")
    lrf   = PatternFill("solid", fgColor="FCE4E4")
    gy    = PatternFill("solid", fgColor="EEEEEE")
    na    = PatternFill("solid", fgColor="F8F8F8")

    # Per-task best speedup over the *S1 model set* — this is the reference for PAR.
    # Using the S1 best ensures S1.PAR-GM and S2.PAR-GM share the same denominator,
    # so S2.PAR-GM directly measures replication of the per-task winner.
    s2_models_only = [m for m in S1 if m in S2]  # exclude V4 Pro (no S2 run)
    best_s1 = {}
    for t in tasks:
        cand = []
        for m in s2_models_only:
            sp = speedup(s1_data[m].get(t))
            if sp: cand.append(sp)
        best_s1[t] = max(cand) if cand else None

    # ---- Sheet 1: Summary ----
    ws = wb.create_sheet("Summary")
    hdr = ["Model",
           "S1 pass", "S2 pass", "Δ pass",
           "S1 GM×", "S2 GM×",
           "S1 PAR-GM", "S2 PAR-GM", "Δ PAR",
           "Paired n", "Paired GM_S1×", "Paired GM_S2×", "GM lift (S2/S1)",
           "S1 fast@10×", "S2 fast@10×",
           "Wins", "Losses", "Ties"]
    ws.append(hdr)
    for c in range(1, len(hdr) + 1):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=1, column=c).alignment = center
        ws.cell(row=1, column=c).fill = blue

    for m in S1:
        s1 = s1_data[m]
        s2 = s2_data.get(m, {})
        s1_sps = [speedup(s1.get(t)) for t in tasks if speedup(s1.get(t))]
        s2_sps = [speedup(s2.get(t)) for t in tasks if speedup(s2.get(t))] if s2 else []
        s1_pass = sum(1 for t in tasks if passed(s1.get(t)))
        s2_pass = sum(1 for t in tasks if passed(s2.get(t))) if s2 else None

        s1_gm = gmean(s1_sps)
        s2_gm = gmean(s2_sps) if s2 else 0
        s1_f10 = sum(1 for x in s1_sps if x >= 10) / n * 100
        s2_f10 = (sum(1 for x in s2_sps if x >= 10) / n * 100) if s2 else None

        # PAR-GM: per-task speedup / per-task best, GM over passed tasks
        s1_par = []
        for t in tasks:
            sp = speedup(s1.get(t))
            if sp and best_s1.get(t):
                s1_par.append(sp / best_s1[t])
        s1_par_gm = gmean(s1_par)
        s2_par = []
        if s2:
            for t in tasks:
                sp = speedup(s2.get(t))
                if sp and best_s1.get(t):
                    s2_par.append(sp / best_s1[t])
        s2_par_gm = gmean(s2_par) if s2 else 0

        # Paired GM (only over tasks where both S1 and S2 passed)
        paired1, paired2, ratios = [], [], []
        wins = losses = ties = 0
        if s2:
            for t in tasks:
                sp1 = speedup(s1.get(t))
                sp2 = speedup(s2.get(t))
                if sp1 and sp2:
                    paired1.append(sp1); paired2.append(sp2); ratios.append(sp2 / sp1)
                    if sp2 > sp1 * 1.10: wins += 1
                    elif sp2 < sp1 * 0.90: losses += 1
                    else: ties += 1
        pg1 = gmean(paired1)
        pg2 = gmean(paired2)
        lift = gmean(ratios) if ratios else 0

        if s2:
            row = [m,
                   s1_pass, s2_pass, s2_pass - s1_pass,
                   s1_gm, s2_gm,
                   s1_par_gm, s2_par_gm, s2_par_gm - s1_par_gm,
                   len(paired1), pg1, pg2, lift,
                   s1_f10, s2_f10,
                   wins, losses, ties]
        else:
            row = [m,
                   s1_pass, "n/a", "n/a",
                   s1_gm, "n/a",
                   s1_par_gm, "n/a", "n/a",
                   "n/a", "n/a", "n/a", "n/a",
                   s1_f10, "n/a",
                   "n/a", "n/a", "n/a"]
        ws.append(row)
        r = ws.max_row
        # speedup × columns
        for col in (5, 6, 11, 12):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, (int, float)):
                ws.cell(row=r, column=col).number_format = '0.00"×"'
        # PAR-GM columns (3 decimals, no × suffix)
        for col in (7, 8, 9):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, (int, float)):
                ws.cell(row=r, column=col).number_format = '0.000'
        # GM lift × column
        ws.cell(row=r, column=13).number_format = '0.00"×"'
        # fast@10×
        for col in (14, 15):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, (int, float)):
                ws.cell(row=r, column=col).number_format = '0.0"%"'
        # color Δ pass
        dp = ws.cell(row=r, column=4).value
        if isinstance(dp, int):
            ws.cell(row=r, column=4).fill = gf if dp > 0 else (rf if dp < 0 else gy)
        # color Δ PAR
        dpar = ws.cell(row=r, column=9).value
        if isinstance(dpar, (int, float)):
            if dpar >= 0.10: ws.cell(row=r, column=9).fill = gf
            elif dpar >= 0.02: ws.cell(row=r, column=9).fill = lgf
            elif dpar <= -0.10: ws.cell(row=r, column=9).fill = rf
            elif dpar <= -0.02: ws.cell(row=r, column=9).fill = lrf
            else: ws.cell(row=r, column=9).fill = gy
        # color GM lift
        lv = ws.cell(row=r, column=13).value
        if isinstance(lv, (int, float)):
            if lv >= 1.5: ws.cell(row=r, column=13).fill = gf
            elif lv >= 1.10: ws.cell(row=r, column=13).fill = lgf
            elif lv <= 0.67: ws.cell(row=r, column=13).fill = rf
            elif lv <= 0.90: ws.cell(row=r, column=13).fill = lrf
            else: ws.cell(row=r, column=13).fill = gy
        # color win/loss/tie
        ws.cell(row=r, column=16).fill = gf
        ws.cell(row=r, column=17).fill = rf
        ws.cell(row=r, column=18).fill = gy
        # gray-out V4 Pro N/A cells (S2-related columns)
        if not s2:
            for col in (3, 4, 6, 8, 9, 10, 11, 12, 13, 15, 16, 17, 18):
                ws.cell(row=r, column=col).fill = na

    ws.column_dimensions["A"].width = 20
    for c in range(2, len(hdr) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "B2"

    # Aggregate row across the 7 S2 models
    all_lifts = []
    for m in S2:
        s1 = s1_data[m]; s2 = s2_data[m]
        for t in tasks:
            sp1 = speedup(s1.get(t)); sp2 = speedup(s2.get(t))
            if sp1 and sp2: all_lifts.append(sp2 / sp1)
    # Aggregate PAR-GM across S1 vs S2 over the 7 S2 models on all paired tasks
    s1_par_pool, s2_par_pool = [], []
    for m in s2_models_only:
        for t in tasks:
            sp1 = speedup(s1_data[m].get(t))
            sp2 = speedup(s2_data[m].get(t))
            if best_s1.get(t):
                if sp1: s1_par_pool.append(sp1 / best_s1[t])
                if sp2: s2_par_pool.append(sp2 / best_s1[t])
    s1_par_agg = gmean(s1_par_pool)
    s2_par_agg = gmean(s2_par_pool)

    ws.append([])
    # Layout (cols): 1=Model, 2=S1pass, 3=S2pass, 4=Δp, 5=S1GM×, 6=S2GM×,
    #                7=S1 PAR-GM, 8=S2 PAR-GM, 9=ΔPAR,
    #                10=pairn, 11=pGM1, 12=pGM2, 13=lift,
    #                14=S1 fast@10×, 15=S2 fast@10×,
    #                16=Wins, 17=Losses, 18=Ties
    ws.append(["Aggregate (7 S2 models, paired)",
               "", "", "", "", "",
               s1_par_agg, s2_par_agg, s2_par_agg - s1_par_agg,
               len(all_lifts), "", "", gmean(all_lifts) if all_lifts else 0,
               "", "",
               sum(1 for x in all_lifts if x>1.10),
               sum(1 for x in all_lifts if x<0.90),
               sum(1 for x in all_lifts if 0.90<=x<=1.10)])
    ar = ws.max_row
    ws.cell(row=ar, column=1).font = bold
    for col in (7, 8, 9):
        ws.cell(row=ar, column=col).number_format = '0.000'
    ws.cell(row=ar, column=13).number_format = '0.00"×"'
    ws.cell(row=ar, column=13).fill = gf
    ws.cell(row=ar, column=9).fill = gf if (s2_par_agg - s1_par_agg) > 0.02 else gy
    ws.cell(row=ar, column=16).fill = gf
    ws.cell(row=ar, column=17).fill = rf
    ws.cell(row=ar, column=18).fill = gy

    # Footer
    ws.append([])
    note_r = ws.max_row + 1
    ws.cell(row=note_r, column=1, value=(
        f"Universe: {n} eligible tasks at {SIZE}. "
        f"GM = geometric mean over passed tasks. "
        f"Paired = tasks where both S1 and S2 passed. "
        f"Lift threshold for Win/Loss = ±10%. "
    )).font = italic
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=len(hdr))

    # ---- Sheet 2: Per-task ----
    ws2 = wb.create_sheet("Per-task")
    h = ["Task"]
    s2_models = list(S2.keys())
    for m in s2_models:
        h += [f"{m} S1", f"{m} S2", f"{m} Δ"]
    ws2.append(h)
    for c in range(1, len(h) + 1):
        ws2.cell(row=1, column=c).font = bold
        ws2.cell(row=1, column=c).alignment = center
        ws2.cell(row=1, column=c).fill = blue
    band_fills = [PatternFill("solid", fgColor=col) for col in
                  ("F2F2F2", "EAF1FB", "FCEAEA", "EAFCEA", "FCFCEA", "F5EAFC", "FCEDEA")]

    for t in tasks:
        row = [t]
        for m in s2_models:
            sp1 = speedup(s1_data[m].get(t))
            sp2 = speedup(s2_data[m].get(t))
            if sp1 is None and sp2 is None:
                row += ["", "", ""]
            else:
                row += [sp1 if sp1 else "—",
                        sp2 if sp2 else "—",
                        (sp2/sp1) if (sp1 and sp2 and sp1>0) else ""]
        ws2.append(row)
        r = ws2.max_row
        for i, m in enumerate(s2_models):
            band = band_fills[i % len(band_fills)]
            cs1, cs2, cd = 2 + i*3, 3 + i*3, 4 + i*3
            for col in (cs1, cs2, cd):
                ws2.cell(row=r, column=col).fill = band
                ws2.cell(row=r, column=col).alignment = center
            for col in (cs1, cs2):
                v = ws2.cell(row=r, column=col).value
                if isinstance(v, (int, float)):
                    ws2.cell(row=r, column=col).number_format = '0.0"×"'
            ratio = ws2.cell(row=r, column=cd).value
            if isinstance(ratio, (int, float)):
                ws2.cell(row=r, column=cd).number_format = '0.00"x"'
                if ratio >= 1.5: ws2.cell(row=r, column=cd).fill = gf
                elif ratio >= 1.10: ws2.cell(row=r, column=cd).fill = lgf
                elif ratio <= 0.67: ws2.cell(row=r, column=cd).fill = rf
                elif ratio <= 0.90: ws2.cell(row=r, column=cd).fill = lrf
                else: ws2.cell(row=r, column=cd).fill = gy

    ws2.freeze_panes = "B2"
    ws2.column_dimensions["A"].width = 32
    for c in range(2, len(h) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 10

    # ---- Sheet 3: README ----
    ws3 = wb.create_sheet("README", index=0)
    ws3["A1"] = "AccelEval Stage 1 (no guidance) vs Stage 2 (with guidance)"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3["A3"] = "Generated:"; ws3["B3"] = datetime.now().isoformat(timespec="seconds")
    ws3["A4"] = "Universe:"; ws3["B4"] = f"{n} tasks at {SIZE}"
    ws3["A5"] = "Headline metric:"; ws3["B5"] = "Geometric mean of speedups over passed tasks"
    ws3["A6"] = "Why GM?:"; ws3["B6"] = "Speedups are multiplicative — GM is invariant to direction (CPU/GPU vs GPU/CPU) and not biased by outliers like arithmetic mean."
    ws3["A8"] = "Sheet 1 — Summary"
    ws3["A9"]  = "  S1 pass / S2 pass: number of tasks compiled+correct."
    ws3["A10"] = "  S1 GM× / S2 GM×: geometric mean of speedups over each model's passed tasks."
    ws3["A11"] = "  S1 PAR-GM / S2 PAR-GM: Peak Attainment Ratio, geometric mean."
    ws3["A12"] = "    PAR_t,m = speedup(t,m) / max_{m' in S1}(speedup(t,m'))   (per-task best of the S1 set)"
    ws3["A13"] = "    PAR-GM = GM of PAR over the model's passed tasks. ∈ [0,1]; 1.0 means model is the per-task winner everywhere."
    ws3["A14"] = "    Reference = S1 best so that S1 and S2 PAR-GM share a denominator and S2 PAR-GM measures replication of the per-task winner."
    ws3["A15"] = "  Δ PAR = S2 PAR-GM - S1 PAR-GM. The headline metric for guidance effectiveness."
    ws3["A16"] = "  Paired GM_S1 / GM_S2: GM over the subset where BOTH S1 and S2 passed."
    ws3["A17"] = "  GM lift (S2/S1): geometric mean of per-task ratios (paired)."
    ws3["A18"] = "  Wins / Losses / Ties: paired tasks where S2 ≥ 1.10× / ≤ 0.90× / between."
    ws3["A20"] = "Sheet 2 — Per-task"
    ws3["A21"] = "  Per task per model: S1 speedup | S2 speedup | S2/S1 ratio."
    ws3["A22"] = "  Δ-cell colors: green ≥1.5×, light-green ≥1.10×, gray neutral, light-red ≤0.90×, red ≤0.67×."
    ws3["A24"] = "Note on V4 Pro:"
    ws3["A25"] = "  DeepSeek V4 Pro was added later; both Stage 1 and Stage 2 are present."
    ws3["A26"] = "  V4 Pro Stage 1 pass count is lower because not all 42 tasks generated successfully."
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 90

    out = ROOT / "runs" / "reports" / f"acceleval_s1_vs_s2_gm_{datetime.now():%Y%m%d_%H%M}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote: {out}")
    print(f"  size: {out.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()

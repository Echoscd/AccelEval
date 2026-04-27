#!/usr/bin/env python3
"""
Build xlsx leaderboard from the canonical acceleval_consolidated_*.json.

Usage:
    python3 scripts/export_xlsx_from_consolidated.py
    python3 scripts/export_xlsx_from_consolidated.py --in <path> --out <path>
"""
import argparse
import glob
import json
import math
import os
import statistics as st
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent


def gmean(xs):
    """Geometric mean over positive values (silently drops <=0/None)."""
    xs = [x for x in xs if x and x > 0]
    if not xs:
        return 0.0
    return math.exp(sum(math.log(x) for x in xs) / len(xs))


def per_task_best(records_size, tasks, models):
    """For each task, the best speedup among all models (None if no model passed)."""
    best = {}
    for t in tasks:
        sps = []
        for m in models:
            rec = records_size[t].get(m)
            if rec and rec.get("compiled") and rec.get("correct"):
                sp = (rec.get("benchmark") or {}).get("speedup_e2e")
                if sp and sp > 0:
                    sps.append(sp)
        best[t] = max(sps) if sps else None
    return best


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="input", default=None,
                    help="consolidated JSON (default: latest in runs/reports/)")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    if args.input:
        in_path = Path(args.input)
    else:
        candidates = sorted(glob.glob(
            str(ROOT / "runs" / "reports" / "acceleval_consolidated_*.json")
        ))
        if not candidates:
            raise SystemExit("no consolidated JSON found; run consolidate_eval_data.py first")
        in_path = Path(candidates[-1])
    print(f"Reading: {in_path.name}")

    doc = json.loads(in_path.read_text())
    models = list(doc["models"].keys())
    tasks = doc["active_tasks"]
    records = doc["records"]
    cpu_b = doc["cpu_baselines_ms"]

    out = Path(args.out) if args.out else ROOT / "runs" / "reports" / (
        in_path.stem.replace("acceleval_consolidated", "acceleval_task_level") + ".xlsx"
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    gf = PatternFill("solid", fgColor="C6EFCE")
    yf = PatternFill("solid", fgColor="FFEB9C")
    rf = PatternFill("solid", fgColor="FFC7CE")
    gy = PatternFill("solid", fgColor="DDDDDD")

    def make_size_sheet(wb, size):
        ws = wb.create_sheet(title=size.capitalize())
        header = ["Task", "CPU (ms)"] + models + ["Max sp", "Best", "# pass"]
        ws.append(header)
        for c in range(1, len(header) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")

        for t in tasks:
            cpu = cpu_b[size].get(t)
            row = [t, cpu]
            best_sp = 0.0; best_model = ""; n_pass = 0
            for m in models:
                rec = records[size][t].get(m)
                if rec is None:
                    row.append("no-gen")
                    continue
                if not rec.get("compiled"):
                    row.append("✗cmp")
                elif not rec.get("correct"):
                    row.append("✗crc")
                else:
                    sp = (rec.get("benchmark") or {}).get("speedup_e2e")
                    n_pass += 1
                    if sp and sp > best_sp:
                        best_sp = sp; best_model = m
                    row.append(sp if sp else 0.0)
            row += [best_sp if best_sp > 0 else None, best_model or None, n_pass]
            ws.append(row)

        for r in range(2, len(tasks) + 2):
            for c in range(3, 3 + len(models)):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if isinstance(v, (int, float)):
                    cell.number_format = '0.00"×"'
                    if v >= 10: cell.fill = gf
                    elif v >= 1: cell.fill = yf
                    else: cell.fill = rf
                elif isinstance(v, str): cell.fill = gy
                cell.alignment = Alignment(horizontal="center")
            maxc = ws.cell(row=r, column=3 + len(models))
            if isinstance(maxc.value, (int, float)) and maxc.value > 0:
                maxc.number_format = '0.00"×"'
            cpuc = ws.cell(row=r, column=2)
            if isinstance(cpuc.value, (int, float)):
                cpuc.number_format = "0.0"

        ws.freeze_panes = "C2"
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 12
        for c in range(3, 3 + len(models)):
            ws.column_dimensions[get_column_letter(c)].width = 13

        # Bottom summary rows: per-model pass count, GM speedup, PAR-GM, fast@10x
        # PAR (Peak Attainment Ratio) = model_speedup / per-task best speedup across
        # this comparison set; PAR-GM aggregates over passed tasks geometrically.
        best_sp = per_task_best(records[size], tasks, models)
        ws.append([])
        base_row = ws.max_row + 1
        labels = [f"# pass / {len(tasks)}", "GM× (pass)", "PAR-GM (pass)", "fast@10×"]
        for i, lbl in enumerate(labels):
            ws.cell(row=base_row + i, column=1, value=lbl).font = Font(bold=True, italic=True)
        for c in range(3, 3 + len(models)):
            m = models[c - 3]
            sps_pass = []
            par_pass = []
            n_pass = 0
            for t in tasks:
                rec = records[size][t].get(m)
                if rec and rec.get("compiled") and rec.get("correct"):
                    n_pass += 1
                    sp = (rec.get("benchmark") or {}).get("speedup_e2e")
                    if sp and sp > 0:
                        sps_pass.append(sp)
                        if best_sp.get(t):
                            par_pass.append(sp / best_sp[t])
            ws.cell(row=base_row,     column=c, value=n_pass).alignment = Alignment(horizontal="center")
            gm = gmean(sps_pass)
            ws.cell(row=base_row + 1, column=c, value=gm).number_format = '0.00"×"'
            ws.cell(row=base_row + 1, column=c).alignment = Alignment(horizontal="center")
            par = gmean(par_pass)
            ws.cell(row=base_row + 2, column=c, value=par).number_format = '0.000'
            ws.cell(row=base_row + 2, column=c).alignment = Alignment(horizontal="center")
            f10 = sum(1 for s in sps_pass if s >= 10) / len(tasks) * 100
            ws.cell(row=base_row + 3, column=c, value=f10).number_format = '0.0"%"'
            ws.cell(row=base_row + 3, column=c).alignment = Alignment(horizontal="center")

    for size in ("small", "medium", "large"):
        make_size_sheet(wb, size)

    # Cross-size Summary sheet — aggregates of the 4 metrics across all 3 sizes
    sums = wb.create_sheet(title="Summary")
    blue_fill = PatternFill("solid", fgColor="D9E1F2")
    sums.cell(row=1, column=1, value="Model").font = Font(bold=True)
    sums.cell(row=1, column=1).fill = blue_fill
    sums.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    SIZES = ("small", "medium", "large")
    METRICS = (
        ("# pass / N",       "0",        "pass"),
        ("GM× (pass)",       '0.00"×"', "gm"),
        ("PAR-GM (pass)",    "0.000",   "par"),
        ("fast@10×",         '0.0"%"',  "f10"),
    )
    # Header row 1 = size groups; header row 2 = metric labels
    col = 2
    size_headers = []
    for sz in SIZES:
        sums.cell(row=1, column=col, value=sz.capitalize()).font = Font(bold=True)
        sums.cell(row=1, column=col).fill = blue_fill
        sums.cell(row=1, column=col).alignment = Alignment(horizontal="center")
        sums.merge_cells(start_row=1, start_column=col,
                         end_row=1,   end_column=col + len(METRICS) - 1)
        size_headers.append(col)
        for j, (lbl, _, _) in enumerate(METRICS):
            c = col + j
            sums.cell(row=2, column=c, value=lbl).font = Font(bold=True)
            sums.cell(row=2, column=c).alignment = Alignment(horizontal="center")
            sums.cell(row=2, column=c).fill = blue_fill
        col += len(METRICS)

    for r_idx, m in enumerate(models, start=3):
        sums.cell(row=r_idx, column=1, value=m).font = Font(bold=True)
        col = 2
        for sz in SIZES:
            best_sp_sz = per_task_best(records[sz], tasks, models)
            sps_pass, par_pass, n_pass = [], [], 0
            for t in tasks:
                rec = records[sz][t].get(m)
                if rec and rec.get("compiled") and rec.get("correct"):
                    n_pass += 1
                    sp = (rec.get("benchmark") or {}).get("speedup_e2e")
                    if sp and sp > 0:
                        sps_pass.append(sp)
                        if best_sp_sz.get(t):
                            par_pass.append(sp / best_sp_sz[t])
            gm = gmean(sps_pass)
            par = gmean(par_pass)
            f10 = sum(1 for s in sps_pass if s >= 10) / len(tasks) * 100
            vals = [n_pass, gm, par, f10]
            for j, (_, fmt, _) in enumerate(METRICS):
                cell = sums.cell(row=r_idx, column=col + j, value=vals[j])
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="center")
                # color PAR cells
                if j == 2 and isinstance(par, (int, float)):
                    if par >= 0.7:   cell.fill = gf
                    elif par >= 0.4: cell.fill = yf
                    elif par >= 0.2: cell.fill = PatternFill("solid", fgColor="FFD8B5")
                    else:            cell.fill = rf
            col += len(METRICS)

    sums.column_dimensions["A"].width = 22
    for c in range(2, 2 + len(SIZES) * len(METRICS)):
        sums.column_dimensions[get_column_letter(c)].width = 14
    sums.freeze_panes = "B3"

    # One README sheet with provenance info
    ws = wb.create_sheet(title="README", index=0)
    ws["A1"] = "AccelEval L3 leaderboard"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Generated at:"; ws["B3"] = doc["generated_at"]
    ws["A4"] = "Source JSON:"; ws["B4"] = in_path.name
    ws["A5"] = "Active tasks:"; ws["B5"] = doc["n_active_tasks"]
    ws["A6"] = "Models:"
    for i, (n, run) in enumerate(doc["models"].items(), start=7):
        ws.cell(row=i, column=2, value=n)
        ws.cell(row=i, column=3, value=run)
    ws["A18"] = "Legend:"
    ws["A19"] = "  ≥10× speedup"; ws["A19"].fill = gf
    ws["A20"] = "  1-10× speedup"; ws["A20"].fill = yf
    ws["A21"] = "  <1× (GPU slower than CPU)"; ws["A21"].fill = rf
    ws["A22"] = "  no-gen / ✗cmp / ✗crc"; ws["A22"].fill = gy
    ws["A24"] = "Bottom summary rows (per size sheet):"
    ws["A25"] = "  # pass / N: tasks compiled+correct out of N active tasks."
    ws["A26"] = "  GM× (pass): geometric mean of speedups over the model's passed tasks."
    ws["A27"] = "  PAR-GM (pass): geometric mean of Peak Attainment Ratios over passed tasks."
    ws["A28"] = "    PAR_t = model_speedup_t / max_{m'}(model'_speedup_t); per-task best is over the displayed model set."
    ws["A29"] = "    PAR-GM ∈ [0, 1]; 1.0 means the model is the per-task winner on every passed task."
    ws["A30"] = "  fast@10×: fraction of N tasks where the model achieves ≥10× speedup."
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 55

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote xlsx: {out}")
    print(f"  size: {out.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()

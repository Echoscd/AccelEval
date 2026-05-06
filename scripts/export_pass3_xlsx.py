#!/usr/bin/env python3
"""Export pass@1 vs pass@3 results to a single xlsx workbook.

Sheets:
  Summary       - per (model, size) GM_p1, GM_p3, PAR_p1, PAR_p3, n_p1, n_p3
  PerTask       - per (task, model, size) speedup_p1, speedup_p3 (long format)
  PassCounts    - small-scale task-pass counts under pass@1 vs pass@3
"""
import json, glob, os, sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
os.chdir(str(ROOT))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Re-load merged data
sys.path.insert(0, "scripts")
from merge_pass3 import PASS2_DIRS, MODELS, SIZES, latest_eval, PASS1_FP

import math
from collections import defaultdict

p1 = json.load(open(PASS1_FP))

best_p1 = defaultdict(dict)
best_p3 = defaultdict(dict)
for sz in SIZES:
    for t, recs in p1["records"][sz].items():
        for m in MODELS:
            r = recs.get(m)
            if not r or not r.get("correct"): continue
            sp = (r.get("benchmark") or {}).get("speedup_e2e")
            if sp is None or sp <= 0: continue
            best_p1[(m, sz)][t] = sp
            best_p3[(m, sz)][t] = sp

for m, run_dir in PASS2_DIRS.items():
    d = latest_eval(run_dir)
    for (t, sid, sz), sp in d.items():
        cur = best_p3[(m, sz)].get(t, 0.0)
        if sp > cur:
            best_p3[(m, sz)][t] = sp

# Compute stars (per-task best across models) for PAR
star_p1 = {sz: {} for sz in SIZES}
star_p3 = {sz: {} for sz in SIZES}
for sz in SIZES:
    for m in MODELS:
        for t, sp in best_p1[(m, sz)].items():
            star_p1[sz][t] = max(star_p1[sz].get(t, 0.0), sp)
        for t, sp in best_p3[(m, sz)].items():
            star_p3[sz][t] = max(star_p3[sz].get(t, 0.0), sp)

# ---------- Build workbook ----------
wb = openpyxl.Workbook()
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
center = Alignment(horizontal="center")
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row=1, ncol=10):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

# Sheet 1: Summary (per (model, size))
ws = wb.active
ws.title = "Summary"
ws.append(["Model", "Size", "GM_pass1", "GM_pass3", "PAR_pass1", "PAR_pass3", "N_pass1", "N_pass3"])
style_header(ws, 1, 8)
for m in MODELS:
    for sz in SIZES:
        tasks_p1 = best_p1[(m, sz)]
        tasks_p3 = best_p3[(m, sz)]
        sps_p1 = list(tasks_p1.values())
        sps_p3 = list(tasks_p3.values())
        gm_p1 = math.exp(sum(math.log(s) for s in sps_p1)/len(sps_p1)) if sps_p1 else None
        gm_p3 = math.exp(sum(math.log(s) for s in sps_p3)/len(sps_p3)) if sps_p3 else None
        par_p1 = math.exp(sum(math.log(tasks_p1[t]/star_p1[sz][t]) for t in tasks_p1 if star_p1[sz].get(t, 0) > 0)/len(tasks_p1)) if tasks_p1 else None
        par_p3 = math.exp(sum(math.log(tasks_p3[t]/star_p3[sz][t]) for t in tasks_p3 if star_p3[sz].get(t, 0) > 0)/len(tasks_p3)) if tasks_p3 else None
        ws.append([m, sz, round(gm_p1, 2) if gm_p1 else "", round(gm_p3, 2) if gm_p3 else "",
                   round(par_p1, 3) if par_p1 else "", round(par_p3, 3) if par_p3 else "",
                   len(sps_p1), len(sps_p3)])

for col in "ABCDEFGH":
    ws.column_dimensions[col].width = 16

# Sheet 2: Per-task long format
ws2 = wb.create_sheet("PerTask")
ws2.append(["Model", "Task", "Size", "Speedup_pass1", "Speedup_pass3", "Improvement_p3_over_p1"])
style_header(ws2, 1, 6)
all_tasks = sorted(set(t for sz in SIZES for t in star_p3[sz]))
for m in MODELS:
    for sz in SIZES:
        for t in all_tasks:
            sp_p1 = best_p1[(m, sz)].get(t)
            sp_p3 = best_p3[(m, sz)].get(t)
            if sp_p1 is None and sp_p3 is None: continue
            improvement = ""
            if sp_p1 and sp_p3:
                improvement = round(sp_p3 / sp_p1, 3)
            ws2.append([m, t, sz,
                        round(sp_p1, 2) if sp_p1 else "",
                        round(sp_p3, 2) if sp_p3 else "",
                        improvement])
for col in "ABCDEF":
    ws2.column_dimensions[col].width = 22

# Sheet 3: Pass counts at small (canonical correctness count)
ws3 = wb.create_sheet("PassCounts")
ws3.append(["Model", "PassCount_pass1", "PassCount_pass3", "Delta"])
style_header(ws3, 1, 4)
for m in MODELS:
    n_p1 = len(best_p1[(m, "small")])
    n_p3 = len(best_p3[(m, "small")])
    ws3.append([m, n_p1, n_p3, n_p3 - n_p1])
for col in "ABCD":
    ws3.column_dimensions[col].width = 22

# Sheet 4: Aggregate cross-model GM at each size
ws4 = wb.create_sheet("CrossModel")
ws4.append(["Size", "GM_pass1_avg", "GM_pass3_avg"])
style_header(ws4, 1, 3)
for sz in SIZES:
    gms_p1 = []
    gms_p3 = []
    for m in MODELS:
        sps_p1 = list(best_p1[(m, sz)].values())
        sps_p3 = list(best_p3[(m, sz)].values())
        if sps_p1:
            gms_p1.append(math.exp(sum(math.log(s) for s in sps_p1)/len(sps_p1)))
        if sps_p3:
            gms_p3.append(math.exp(sum(math.log(s) for s in sps_p3)/len(sps_p3)))
    avg_p1 = sum(gms_p1)/len(gms_p1) if gms_p1 else None
    avg_p3 = sum(gms_p3)/len(gms_p3) if gms_p3 else None
    ws4.append([sz, round(avg_p1, 2) if avg_p1 else "", round(avg_p3, 2) if avg_p3 else ""])
for col in "ABC":
    ws4.column_dimensions[col].width = 18

# Save
import datetime
ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
out_dir = ROOT / "runs" / "reports"
out_dir.mkdir(exist_ok=True)
out_fp = out_dir / f"acceleval_pass3_{ts}.xlsx"
wb.save(out_fp)
print(f"Saved: {out_fp}", file=sys.stderr)
print(f"Sheets: Summary, PerTask, PassCounts, CrossModel")
